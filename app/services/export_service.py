"""
Export service for generating Excel files and handling data exports
"""

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from .timesheet_service import TimesheetService


class ExportService:
    """Service class for export operations"""
    
    @staticmethod
    def generate_excel_export(start_date, end_date, user_id):
        """Generate Excel file with time tracking data"""
        # Get the data
        data = TimesheetService.get_export_data(start_date, end_date, user_id)
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Timesheet"
        
        # Generate date range
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
        
        dates = []
        current_date = start_dt
        while current_date <= end_dt:
            dates.append(current_date.strftime('%Y-%m-%d'))
            current_date += timedelta(days=1)
        
        # Create header row
        headers = ['Project ID', 'Project Name'] + dates
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Organize data by project
        project_data = {}
        for row in data:
            project_id = row['project_id']
            project_name = row['project_name']
            date = row['date']
            minutes = row['total_minutes']
            
            if project_id not in project_data:
                project_data[project_id] = {
                    'name': project_name,
                    'dates': {}
                }
            
            project_data[project_id]['dates'][date] = minutes
        
        # Fill in the data
        row_num = 2
        for project_id in sorted(project_data.keys()):
            project_info = project_data[project_id]
            
            # Project ID column
            ws.cell(row=row_num, column=1, value=project_id)
            
            # Project Name column  
            ws.cell(row=row_num, column=2, value=project_info['name'])
            
            # Time entries for each date
            for col, date in enumerate(dates, 3):
                minutes = project_info['dates'].get(date, 0)
                # Convert to hours and only show if there's time tracked
                if minutes > 0:
                    hours = round(minutes / 60, 2)
                    cell = ws.cell(row=row_num, column=col, value=hours)
                    cell.alignment = Alignment(horizontal='right')
                else:
                    # Empty cell for days with no time
                    ws.cell(row=row_num, column=col, value='')
            
            row_num += 1
        
        # Add totals row only if there's data
        if project_data:
            ws.cell(row=row_num, column=1, value='TOTAL')
            ws.cell(row=row_num, column=2, value='All Projects')
            
            # Calculate totals for each date
            for col, date in enumerate(dates, 3):
                total_minutes = sum(
                    project_info['dates'].get(date, 0) 
                    for project_info in project_data.values()
                )
                if total_minutes > 0:
                    total_hours = round(total_minutes / 60, 2)
                    cell = ws.cell(row=row_num, column=col, value=total_hours)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
                    cell.alignment = Alignment(horizontal='right')
                else:
                    # Empty cell for totals when no data
                    cell = ws.cell(row=row_num, column=col, value='')
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
